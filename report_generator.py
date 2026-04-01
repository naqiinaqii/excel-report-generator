import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

# ---- SETTINGS ----
INPUT_FILE = "raw_data.csv"
OUTPUT_FILE = f"report_{datetime.today().strftime('%Y-%m-%d')}.xlsx"
REPORT_TITLE = "Staff Performance Report"

# ---- STEP 1: Load and clean data ----
def load_and_clean(path):
    df = pd.read_csv(path)

    # Standardise text columns — fix inconsistent capitalisation
    df["Status"] = df["Status"].str.strip().str.capitalize()
    df["Department"] = df["Department"].str.strip().str.title()
    df["Name"] = df["Name"].str.strip().str.title()

    # Fill any missing values
    df.fillna("N/A", inplace=True)

    print(f"✅ Loaded {len(df)} records from {path}")
    return df

# ---- STEP 2: Generate summary stats ----
def generate_summary(df):
    summary = {
        "Total Staff": len(df),
        "Active Staff": len(df[df["Status"] == "Active"]),
        "Inactive Staff": len(df[df["Status"] == "Inactive"]),
        "Avg Tasks Completed": round(df["Tasks Completed"].mean(), 1),
        "Avg Hours Worked": round(df["Hours Worked"].mean(), 1),
        "Top Performer": df.loc[df["Tasks Completed"].idxmax(), "Name"]
    }
    return summary

# ---- STEP 3: Write Excel report ----
def write_excel(df, summary, output_path, title):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # --- Styles ---
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    subheader_fill = PatternFill("solid", fgColor="2E75B6")  # medium blue
    title_font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    summary_font = Font(name="Calibri", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # --- Title ---
    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generated on: {datetime.today().strftime('%d %B %Y')}"
    ws["A2"].alignment = center
    ws["A2"].font = Font(italic=True, color="888888")

    # --- Summary Section ---
    ws["A4"] = "SUMMARY"
    ws["A4"].font = Font(bold=True, size=12, color="1F4E79")

    row = 5
    for key, value in summary.items():
        ws.cell(row=row, column=1, value=key).font = summary_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    # --- Data Table Header ---
    table_start_row = row + 2
    ws.cell(row=table_start_row, column=1, value="DETAILED DATA").font = Font(bold=True, size=12, color="1F4E79")
    table_start_row += 1

    columns = list(df.columns)
    for col_num, col_name in enumerate(columns, 1):
        cell = ws.cell(row=table_start_row, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # --- Data Rows ---
    for row_num, row_data in enumerate(df.itertuples(index=False), table_start_row + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            # Highlight inactive staff in light red
            if col_num == len(columns) and value == "Inactive":
                cell.fill = PatternFill("solid", fgColor="FFD7D7")

    # --- Auto column width ---
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    # --- Bar Chart: Tasks Completed per Person ---
    chart_sheet = wb.create_sheet("Chart")

    # Write data for chart
    chart_sheet["A1"] = "Name"
    chart_sheet["B1"] = "Tasks Completed"
    for i, (name, tasks) in enumerate(zip(df["Name"], df["Tasks Completed"]), 2):
        chart_sheet[f"A{i}"] = name
        chart_sheet[f"B{i}"] = tasks

    chart = BarChart()
    chart.type = "col"
    chart.title = "Tasks Completed per Staff"
    chart.y_axis.title = "Tasks"
    chart.x_axis.title = "Staff"
    chart.style = 10

    data_ref = Reference(chart_sheet, min_col=2, min_row=1, max_row=len(df) + 1)
    cats_ref = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(df) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    chart.width = 20
    chart.height = 12

    chart_sheet.add_chart(chart, "D2")

    wb.save(output_path)
    print(f"💾 Report saved as: {output_path}")

# ---- MAIN ----
def main():
    print("🚀 Starting report generation...\n")
    df = load_and_clean(INPUT_FILE)
    summary = generate_summary(df)

    print("\n📊 Summary:")
    for k, v in summary.items():
        print(f"   {k}: {v}")

    print("\n📝 Writing Excel report...")
    write_excel(df, summary, OUTPUT_FILE, REPORT_TITLE)
    print("\n✅ Done! Open your report to see the result.")

main()